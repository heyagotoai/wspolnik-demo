"""Testy importu wpłat z Excela (/api/import/payments)."""

import io

import pytest
from fastapi.testclient import TestClient


def _make_payments_xlsx(
    data_rows: list[tuple],
    *,
    sheet_title: str = "Dopasowania",
    include_nazwisko: bool = False,
) -> bytes:
    """Wiersze: (lokal, data, kwota) lub z nazwiskiem (lokal, nazwisko, data, kwota)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    if include_nazwisko:
        headers = ["Lokal", "Nazwisko", "Data wpłaty", "Kwota"]
    else:
        headers = ["Lokal", "Data wpłaty", "Kwota"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)

    for r_idx, row in enumerate(data_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


APT_1 = {"id": "apt-1", "number": "1", "billing_group_id": None}
APT_2 = {"id": "apt-2", "number": "2", "billing_group_id": None}
APT_47 = {"id": "apt-47", "number": "47", "billing_group_id": None}


class TestPaymentsTemplate:
    def test_admin_pobiera_szablon(self, admin_client: TestClient):
        res = admin_client.get("/api/import/payments-template")
        assert res.status_code == 200
        assert res.content[:4] == b"PK\x03\x04"

    def test_mieszkaniec_brak_dostepu(self, resident_client: TestClient):
        res = resident_client.get("/api/import/payments-template")
        assert res.status_code == 403


class TestImportPaymentsDryRun:
    def test_jeden_lokal(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APT_1])
        fake_sb.set_table_data("charges", [])

        xlsx = _make_payments_xlsx([("1", "08.01.2026", "139,20")])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        data = res.json()
        assert data["updated"] == 1
        assert data["errors"] == 0

    def test_plik_z_nazwiskiem_kolumna_ignorowana(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APT_1])
        fake_sb.set_table_data("charges", [])

        xlsx = _make_payments_xlsx(
            [("1", "KOWALSKI", "08.01.2026", "139,20")],
            include_nazwisko=True,
        )
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        assert res.json()["updated"] == 1

    def test_brak_daty_skip(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APT_1])
        xlsx = _make_payments_xlsx([("1", None, "100,00")])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        d = res.json()
        assert d["skipped"] >= 1

    def test_nieznany_lokal_error(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APT_1])
        xlsx = _make_payments_xlsx([("999", "08.01.2026", "10,00")])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        assert res.json()["errors"] == 1

    def test_dwie_daty_jedna_kwota_ta_sama_kwota_na_kazda_date(self, admin_client: TestClient, fake_sb):
        """Wiele dni księgowania w jednej komórce — jedna kwota = osobna wpłata o tej kwocie na każdą datę."""
        fake_sb.set_table_data("apartments", [APT_47])
        fake_sb.set_table_data("charges", [])
        xlsx = _make_payments_xlsx([
            ("47", "10.02.2026; 27.02.2026", "341,20"),
        ])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        d = res.json()
        assert d["updated"] == 2
        assert d["errors"] == 0
        msgs = [r.get("message") or "" for r in d["rows"]]
        assert any("(1/2)" in m for m in msgs)
        assert any("(2/2)" in m for m in msgs)

    def test_dwie_daty_dwie_kwoty_jawnie(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APT_47])
        fake_sb.set_table_data("charges", [])
        xlsx = _make_payments_xlsx([
            ("47", "10.02.2026; 27.02.2026", "341,20; 341,20"),
        ])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        assert res.json()["updated"] == 2
        assert res.json()["errors"] == 0

    def test_dwie_daty_wiele_lokali_blad(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APT_1, APT_2])
        xlsx = _make_payments_xlsx([
            ("1,2", "10.02.2026; 27.02.2026", "100,00"),
        ])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        assert res.json()["errors"] == 1

    def test_dedup_istniejaca_wplata_w_bazi(self, admin_client: TestClient, fake_sb):
        """Ta sama para (lokal, data) co w payments → pominięcie, jak w imporcie z banku."""
        fake_sb.set_table_data("apartments", [APT_1])
        fake_sb.set_table_data("charges", [])
        fake_sb.set_table_data("payments", [
            {"apartment_id": "apt-1", "payment_date": "2026-01-08", "amount": "50.00"},
        ])
        xlsx = _make_payments_xlsx([("1", "08.01.2026", "139,20")])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        d = res.json()
        assert d["updated"] == 0
        assert d["skipped"] == 1
        assert "duplikat" in (d["rows"][0].get("message") or "").lower()

    def test_dedup_drugi_wiersz_pliku_ta_sama_data(self, admin_client: TestClient, fake_sb):
        """W jednym imporcie drugi wiersz z tą samą datą/lokalem jest pomijany."""
        fake_sb.set_table_data("apartments", [APT_1])
        fake_sb.set_table_data("charges", [])
        fake_sb.set_table_data("payments", [])
        xlsx = _make_payments_xlsx([
            ("1", "08.01.2026", "100,00"),
            ("1", "08.01.2026", "200,00"),
        ])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        d = res.json()
        assert d["updated"] == 1
        assert d["skipped"] == 1

    def test_dedup_wiele_dat_czesc_z_bazy(self, admin_client: TestClient, fake_sb):
        """Przy wielu datach w komórce tylko kolizyjne są pomijane, reszta się importuje."""
        fake_sb.set_table_data("apartments", [APT_47])
        fake_sb.set_table_data("charges", [])
        fake_sb.set_table_data("payments", [
            {"apartment_id": "apt-47", "payment_date": "2026-02-10", "amount": "1.00"},
        ])
        xlsx = _make_payments_xlsx([
            ("47", "10.02.2026; 27.02.2026", "341,20"),
        ])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        d = res.json()
        assert d["updated"] == 1
        assert d["skipped"] == 1

    def test_dedup_import_zbiorczy_gdy_lokal_ma_juz_wplate(self, admin_client: TestClient, fake_sb):
        """Przy wierszu wielolokalowym pomijamy całość, jeśli którykolwiek lokal ma wpłatę w tym dniu."""
        fake_sb.set_table_data("apartments", [APT_1, APT_2])
        fake_sb.set_table_data("charges", [])
        fake_sb.set_table_data("payments", [
            {"apartment_id": "apt-1", "payment_date": "2026-02-10", "amount": "10.00"},
        ])
        xlsx = _make_payments_xlsx([("1,2", "10.02.2026", "100,00")])
        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 200
        d = res.json()
        assert d["updated"] == 0
        assert d["skipped"] == 1
        assert "zbiorcza" in (d["rows"][0].get("message") or "").lower()

    def test_brak_wymaganych_kolumn_422(self, admin_client: TestClient, fake_sb):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Dopasowania"
        ws.cell(row=1, column=1, value="Lokal")
        ws.cell(row=1, column=2, value="Kwota")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        res = admin_client.post(
            "/api/import/payments?dry_run=true",
            files={"file": ("w.xlsx", buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 422
