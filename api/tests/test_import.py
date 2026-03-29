"""Testy endpointów /api/import.

Pokryte scenariusze:
- Pobieranie szablonu (GET /import/template)
- Import poprawnych danych (dry_run=True i False)
- Grupa lokali w jednej komórce (przecinki, „25.26”) → to samo saldo dla każdego; częściowy brak numerów → updated + komunikat
- Wiersz z nieznanym numerem lokalu → skipped
- Wiersz z nieprawidłowym saldem → error
- Plik bez wiersza data_salda → HTTP 422
- Nieprawidłowa data_salda → HTTP 422
- Plik bez nagłówków → HTTP 422
- Plik .txt zamiast .xlsx → HTTP 422
- Kilka lokali z różnymi statusami jednocześnie
- Różne formaty daty (YYYY-MM-DD, DD.MM.YYYY)
"""

import io

import pytest
from fastapi.testclient import TestClient


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _make_xlsx(
    data_rows: list[dict],
    balance_date: str = "2024-12-31",
    include_date_row: bool = True,
    include_header_row: bool = True,
    date_label: str = "data_salda",
) -> bytes:
    """
    Generuj plik .xlsx w formacie wymaganym przez import:
      Wiersz 1: | data_salda | <date> |
      Wiersz 2: | numer_lokalu | saldo_poczatkowe |
      Wiersz 3+: dane
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    row_offset = 1

    if include_date_row:
        ws.cell(row=row_offset, column=1, value=date_label)
        ws.cell(row=row_offset, column=2, value=balance_date)
        row_offset += 1

    if include_header_row:
        ws.cell(row=row_offset, column=1, value="numer_lokalu")
        ws.cell(row=row_offset, column=2, value="saldo_poczatkowe")
        row_offset += 1

    for row in data_rows:
        ws.cell(row=row_offset, column=1, value=row.get("numer_lokalu"))
        ws.cell(row=row_offset, column=2, value=row.get("saldo_poczatkowe"))
        row_offset += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


APARTMENT_1 = {"id": "apt-1", "number": "1A"}
APARTMENT_2 = {"id": "apt-2", "number": "2B"}


# ──────────────────────────────────────────────────────────────────────────────
# Pobieranie szablonu
# ──────────────────────────────────────────────────────────────────────────────

class TestDownloadTemplate:
    def test_zwraca_plik_xlsx(self, admin_client: TestClient):
        res = admin_client.get("/api/import/template")
        assert res.status_code == 200
        assert "spreadsheetml" in res.headers["content-type"]
        assert res.content[:4] == b"PK\x03\x04"  # sygnatura .zip/.xlsx

    def test_zawartosc_szablonu_zawiera_date_i_kolumny(self, admin_client: TestClient):
        from openpyxl import load_workbook
        res = admin_client.get("/api/import/template")
        wb = load_workbook(io.BytesIO(res.content))
        ws = wb.active
        # Wiersz 1: data_salda
        assert str(ws.cell(row=1, column=1).value).lower() == "data_salda"
        # Wiersz 2: nagłówki
        headers = [ws.cell(row=2, column=c).value for c in range(1, 3)]
        assert "numer_lokalu" in headers
        assert "saldo_poczatkowe" in headers
        # Brak kolumny email_wlasciciela
        all_headers = [ws.cell(row=2, column=c).value for c in range(1, 5)]
        assert "email_wlasciciela" not in all_headers

    def test_mieszkaniec_nie_ma_dostepu(self, resident_client: TestClient):
        res = resident_client.get("/api/import/template")
        assert res.status_code == 403


# ──────────────────────────────────────────────────────────────────────────────
# Import — walidacja formatu pliku
# ──────────────────────────────────────────────────────────────────────────────

class TestImportValidation:
    def _upload(self, client: TestClient, content: bytes, filename: str = "test.xlsx"):
        return client.post(
            "/api/import/initial-state?dry_run=true",
            files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    def test_plik_txt_zwraca_422(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])
        res = admin_client.post(
            "/api/import/initial-state?dry_run=true",
            files={"file": ("test.txt", b"cos", "text/plain")},
        )
        assert res.status_code == 422
        assert ".xlsx" in res.json()["detail"]

    def test_brak_wiersza_data_salda_zwraca_422(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])
        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": 0}], include_date_row=False)
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 422
        assert "data_salda" in res.json()["detail"]

    def test_nieprawidlowa_data_salda_zwraca_422(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])
        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": 0}], balance_date="nie-data")
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 422
        assert "data_salda" in res.json()["detail"]

    def test_brak_naglowkow_zwraca_422(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])
        xlsx = _make_xlsx([], include_header_row=False)
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 422
        assert "nagłówk" in res.json()["detail"]

    def test_mieszkaniec_nie_ma_dostepu(self, resident_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])
        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": 0}])
        res = resident_client.post(
            "/api/import/initial-state?dry_run=true",
            files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert res.status_code == 403


# ──────────────────────────────────────────────────────────────────────────────
# Import — dry_run (podgląd)
# ──────────────────────────────────────────────────────────────────────────────

class TestImportDryRun:
    def _upload(self, client: TestClient, content: bytes):
        return client.post(
            "/api/import/initial-state?dry_run=true",
            files={"file": ("test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    def test_poprawny_import_dry_run(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": -200.50}])
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        data = res.json()
        assert data["dry_run"] is True
        assert data["updated"] == 1
        assert data["skipped"] == 0
        assert data["errors"] == 0
        assert data["rows"][0]["status"] == "updated"
        assert data["rows"][0]["apartment_number"] == "1A"

    def test_nieznany_lokal_jest_pomijany(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([{"numer_lokalu": "99X", "saldo_poczatkowe": 0}])
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        data = res.json()
        assert data["updated"] == 0
        assert data["skipped"] == 1
        assert data["rows"][0]["status"] == "skipped"

    def test_nieprawidlowe_saldo_zwraca_blad_wiersza(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": "nie_liczba"}])
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        data = res.json()
        assert data["errors"] == 1
        assert data["rows"][0]["status"] == "error"
        assert "saldo" in data["rows"][0]["message"].lower()

    def test_kilka_lokali_rozne_statusy(self, admin_client: TestClient, fake_sb):
        """1 poprawny + 1 nieznany + 1 zły format → updated=1, skipped=1, errors=1."""
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([
            {"numer_lokalu": "1A", "saldo_poczatkowe": -100},
            {"numer_lokalu": "99X", "saldo_poczatkowe": 0},
            {"numer_lokalu": "1A", "saldo_poczatkowe": "abc"},
        ])
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        data = res.json()
        assert data["updated"] == 1
        assert data["skipped"] == 1
        assert data["errors"] == 1

    def test_brak_numeru_lokalu_skipuje_wiersz(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([
            {"numer_lokalu": None, "saldo_poczatkowe": 0},
            {"numer_lokalu": "1A", "saldo_poczatkowe": 50},
        ])
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        assert res.json()["updated"] == 1
        assert res.json()["skipped"] == 1


# ──────────────────────────────────────────────────────────────────────────────
# Import — dry_run=False (zapis)
# ──────────────────────────────────────────────────────────────────────────────

class TestImportApply:
    def _upload(self, client: TestClient, content: bytes):
        return client.post(
            "/api/import/initial-state?dry_run=false",
            files={"file": ("test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    def test_import_zapisuje_dane(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": -200.50}])
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        data = res.json()
        assert data["dry_run"] is False
        assert data["updated"] == 1
        assert data["errors"] == 0

    def test_data_format_polski(self, admin_client: TestClient, fake_sb):
        """Format DD.MM.YYYY powinien być poprawnie parsowany."""
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": 50}], balance_date="31.12.2024")
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        assert res.json()["updated"] == 1
        assert res.json()["errors"] == 0

    def test_saldo_zero_jest_akceptowane(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": 0}])
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        assert res.json()["updated"] == 1

    def test_wiele_lokali_naraz(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1, APARTMENT_2])

        xlsx = _make_xlsx([
            {"numer_lokalu": "1A", "saldo_poczatkowe": -100},
            {"numer_lokalu": "2B", "saldo_poczatkowe": 200},
        ])
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        assert res.json()["updated"] == 2
        assert res.json()["errors"] == 0

    def test_data_iso_format(self, admin_client: TestClient, fake_sb):
        """Format YYYY-MM-DD powinien być poprawnie parsowany."""
        fake_sb.set_table_data("apartments", [APARTMENT_1])

        xlsx = _make_xlsx([{"numer_lokalu": "1A", "saldo_poczatkowe": -50}], balance_date="2024-12-31")
        res = self._upload(admin_client, xlsx)

        assert res.status_code == 200
        assert res.json()["updated"] == 1

APARTMENT_25 = {"id": "apt-25", "number": "25"}
APARTMENT_26 = {"id": "apt-26", "number": "26"}
APARTMENT_32 = {"id": "apt-32", "number": "32"}
APARTMENT_45 = {"id": "apt-45", "number": "45"}


class TestImportGrupyLokali:
    """Grupy w jednej komórce: przecinki lub zapis 25.26 (dwie liczby 2+ cyfry)."""

    def _upload(self, client: TestClient, content: bytes):
        return client.post(
            "/api/import/initial-state?dry_run=true",
            files={"file": ("test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    def test_grupa_po_przecinku(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data(
            "apartments",
            [
                {"id": "apt-18", "number": "18"},
                {"id": "apt-31", "number": "31"},
                {"id": "apt-42", "number": "42"},
                {"id": "apt-44", "number": "44"},
            ],
        )
        xlsx = _make_xlsx([{"numer_lokalu": "18,31,42,44", "saldo_poczatkowe": -10}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        data = res.json()
        assert data["updated"] == 1
        assert data["skipped"] == 0
        assert data["rows"][0]["status"] == "updated"
        assert data["rows"][0]["message"] is None

    def test_grupa_kropka_dwucyfrowe(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_25, APARTMENT_26])
        xlsx = _make_xlsx([{"numer_lokalu": "25.26", "saldo_poczatkowe": 100}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        data = res.json()
        assert data["updated"] == 1
        assert data["rows"][0]["status"] == "updated"

    def test_grupa_kropka_32_45(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_32, APARTMENT_45])
        xlsx = _make_xlsx([{"numer_lokalu": "32.45", "saldo_poczatkowe": -50}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        assert res.json()["rows"][0]["status"] == "updated"

    def test_grupa_czesciowo_nieistniejace_ostrzezenie(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_1])
        xlsx = _make_xlsx([{"numer_lokalu": "1A,99X", "saldo_poczatkowe": 20}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        data = res.json()
        assert data["updated"] == 1
        row = data["rows"][0]
        assert row["status"] == "updated"
        assert row["message"] is not None
        assert "99X" in row["message"]

    def test_grupa_kropka_jako_float_excel(self, admin_client: TestClient, fake_sb):
        """Excel zapisuje 32.45 jako float — musi rozpoznać dwa lokale (32 i 45)."""
        fake_sb.set_table_data("apartments", [APARTMENT_32, APARTMENT_45])
        xlsx = _make_xlsx([{"numer_lokalu": 32.45, "saldo_poczatkowe": -50}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        data = res.json()
        assert data["updated"] == 1
        assert data["rows"][0]["status"] == "updated"

    def test_grupa_25_26_jako_float_excel(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data("apartments", [APARTMENT_25, APARTMENT_26])
        xlsx = _make_xlsx([{"numer_lokalu": 25.26, "saldo_poczatkowe": 100}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        assert res.json()["updated"] == 1
        assert res.json()["rows"][0]["status"] == "updated"

    def test_grupa_3_4a_tekst_z_kropka(self, admin_client: TestClient, fake_sb):
        """Tekst 3.4A (czesto Excel zapisuje kropke zamiast przecinka) -> lokale 3 i 4A."""
        fake_sb.set_table_data(
            "apartments",
            [{"id": "a3", "number": "3"}, {"id": "a4a", "number": "4A"}],
        )
        xlsx = _make_xlsx([{"numer_lokalu": "3.4A", "saldo_poczatkowe": 5}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        assert res.json()["updated"] == 1
        assert res.json()["rows"][0]["status"] == "updated"

    def test_grupa_3_4a_przecinek_tekst(self, admin_client: TestClient, fake_sb):
        fake_sb.set_table_data(
            "apartments",
            [{"id": "a3", "number": "3"}, {"id": "a4a", "number": "4A"}],
        )
        xlsx = _make_xlsx([{"numer_lokalu": "3,4A", "saldo_poczatkowe": 5}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        assert res.json()["updated"] == 1

    def test_lokal_zbiorczy_3_4a_jeden_rekord(self, admin_client: TestClient, fake_sb):
        """W bazie jeden lokal „3,4A” — import bez rozbijania na 3 i 4A."""
        fake_sb.set_table_data("apartments", [{"id": "apt-34a", "number": "3,4A"}])
        xlsx = _make_xlsx([{"numer_lokalu": "3,4A", "saldo_poczatkowe": -100}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        data = res.json()
        assert data["updated"] == 1
        assert data["skipped"] == 0
        assert data["rows"][0]["status"] == "updated"

    def test_lokal_zbiorczy_25_26_jeden_rekord(self, admin_client: TestClient, fake_sb):
        """W bazie jeden lokal „25,26” — nie szukamy osobno 25 i 26."""
        fake_sb.set_table_data("apartments", [{"id": "apt-2526", "number": "25,26"}])
        xlsx = _make_xlsx([{"numer_lokalu": "25,26", "saldo_poczatkowe": 0}])
        res = self._upload(admin_client, xlsx)
        assert res.status_code == 200
        assert res.json()["updated"] == 1
        assert res.json()["rows"][0]["status"] == "updated"

