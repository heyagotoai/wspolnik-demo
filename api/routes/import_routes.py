"""Import z Excel: stan początkowy (saldo), wpłaty z arkusza dopasowań, zestawienia bankowe."""

import io
import logging
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse

from api.core.payment_split import compute_split_amounts
from api.core.security import require_admin
from api.core.supabase_client import get_supabase
from api.models.schemas import (
    ImportInitialStateResult,
    ImportPaymentsResult,
    ImportRowResult,
    ImportBankStatementResult,
    BankStatementMatchedRow,
    BankStatementUnmatchedRow,
)
from api.services.bank_statement_parser import (
    ApartmentRecord,
    UnmatchedTransaction,
    parse_bank_statement,
)

router = APIRouter(prefix="/import", tags=["import"])
logger = logging.getLogger(__name__)

MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB

# ──────────────────────────────────────────────
# Format pliku:
#   Wiersz 1:   | data_salda | <data> |
#   Wiersz 2:   | numer_lokalu | saldo_poczatkowe |
#   Wiersz 3+:  dane lokali
# ──────────────────────────────────────────────


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────

def _parse_date(value) -> date | None:
    """Parsuj datę z formatu YYYY-MM-DD lub DD.MM.YYYY. Obsługuje też obiekty date/datetime z openpyxl."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value) -> Decimal | None:
    """Parsuj liczbę dziesiętną z dowolnego formatu (przecinek/kropka jako separator)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip().replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _normalize_str(value) -> str:
    """Zamień wartość komórki na string i przytnij białe znaki."""
    if value is None:
        return ""
    return str(value).strip()



def _normalize_apartment_separators(s: str) -> str:
    """Ujednolicenie przecinka/kropki z Excela (w tym znaki „wide\" U+FF0C / U+FF0E) oraz NBSP."""
    if not s:
        return ""
    for code in (0xFF0C, 0xFE50, 0x201A, 0x060C, 0x3001):
        s = s.replace(chr(code), ",")
    for code in (0xFF0E, 0x2024):
        s = s.replace(chr(code), ".")
    s = s.replace(chr(0x00A0), " ")
    return s.strip()


def _cell_to_apartment_text(value) -> str:
    """
    Tekst z komórki numer_lokalu.

    Excel (PL) pokazuje przecinki, ale komórka liczbowa jest w pamięci jako float — Python
    zapisuje ją z kropką (np. 25,26 z arkusza → 25.26). To nie zmienia znaczenia: dalej
    rozbijamy na lokale 25 i 26. Tekst „3,4A” zostaje z przecinkami do _parse_apartment_numbers.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):
            return ""
        r = round(value, 6)
        if abs(r - round(r)) < 1e-9:
            return str(int(round(r)))
        return format(r, ".6f").rstrip("0").rstrip(".")
    if isinstance(value, (datetime, date)):
        return ""
    return _normalize_apartment_separators(str(value).strip())


def _expand_apartment_chunk(chunk: str) -> list[str]:
    """Rozbija fragment (np. „25.26”, „3.4A”) po zasadach Excel/PL."""
    if re.fullmatch(r"\d{2,}\.\d{2,}", chunk):
        left, right = chunk.split(".", 1)
        return [left, right]
    m = re.fullmatch(r"(\d+)\.(\d+[A-Za-z]+)", chunk)
    if m:
        return [m.group(1), m.group(2)]
    return [chunk]


def _parse_apartment_numbers(raw: str) -> list[str]:
    """
    Rozbija komórkę numer_lokalu na pojedyncze numery.

    - Lista po przecinku/średniku/tabulatorze/pionowej kresce: „18,31,42”, „3,4A”, „25,26”.
    - Gdy Excel zwrócił float (np. 25,26 jako liczba), w raw jest „25.26” — rozbicie po kropce
      na dwie grupy cyfr (≥2 cyfry z każdej strony), żeby uzyskać lokale 25 i 26.
    - Gdy Excel zapisał „3,4A” jako tekst „3.4A” (kropka zamiast przecinka) — rozbicie na 3 i 4A.
    """
    s = _normalize_apartment_separators(_normalize_str(raw))
    if not s:
        return []
    out: list[str] = []
    for chunk in re.split("[,;	|]+", s):
        chunk = chunk.strip()
        if not chunk:
            continue
        out.extend(_expand_apartment_chunk(chunk))
    return out

def _build_template_xlsx() -> bytes:
    """Generuj plik szablonu .xlsx."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="Biblioteka openpyxl nie jest zainstalowana")

    wb = Workbook()
    ws = wb.active
    ws.title = "Stan początkowy"

    date_label_fill = PatternFill("solid", fgColor="FFF2CC")
    date_label_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    # Wiersz 1: data_salda (global)
    ws.cell(row=1, column=1, value="data_salda").font = date_label_font
    ws.cell(row=1, column=1).fill = date_label_fill
    ws.cell(row=1, column=2, value="2024-12-31").fill = date_label_fill
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="left")

    # Wiersz 2: nagłówki kolumn
    for col, header in enumerate(["numer_lokalu", "saldo_poczatkowe"], start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Przykładowe wiersze
    examples = [
        ("1A", -200.50),
        ("1B", 0.00),
        ("2A", 150.00),
        ("3A", -50.75),
    ]
    for row_idx, (nr, saldo) in enumerate(examples, start=3):
        ws.cell(row=row_idx, column=1, value=nr)
        ws.cell(row=row_idx, column=2, value=saldo)

    # Szerokości kolumn
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22

    # Komentarz
    ws.cell(row=9, column=1, value="Uwagi:").font = Font(bold=True)
    ws.cell(row=10, column=1, value="• data_salda (wiersz 1, kolumna B): format YYYY-MM-DD lub DD.MM.YYYY — obowiązuje dla wszystkich lokali")
    ws.cell(row=11, column=1, value="• numer_lokalu musi odpowiadać istniejącemu lokalowi w systemie")
    ws.cell(row=12, column=1, value="• grupa lokali w jednej komórce: oddziel przecinkiem lub średnikiem; zapis „25.26” (dwie liczby po 2+ cyfr) = lokale 25 i 26")
    ws.cell(row=13, column=1, value="• saldo_poczatkowe: ujemne = zaległość, dodatnie = nadpłata (PLN) — wspólne dla wszystkich lokali z wiersza")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _parse_workbook(content: bytes):
    """
    Parsuj plik xlsx i zwróć (global_date, col_idx, data_rows, header_row_idx).

    Oczekiwany format:
      Wiersz 1: | data_salda | <data> |
      Wiersz 2: | numer_lokalu | saldo_poczatkowe |
      Wiersz 3+: dane
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Biblioteka openpyxl nie jest zainstalowana")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Nie można otworzyć pliku Excel: {e}")

    if not all_rows:
        raise HTTPException(status_code=422, detail="Plik jest pusty")

    # Szukaj wiersza data_salda i wiersza nagłówków
    global_date: date | None = None
    header_row_idx: int | None = None

    for i, row in enumerate(all_rows):
        first = _normalize_str(row[0] if row else None).lower()
        if first == "data_salda":
            val = row[1] if len(row) > 1 else None
            global_date = _parse_date(val)
            if global_date is None:
                raise HTTPException(
                    status_code=422,
                    detail=f"Nieprawidłowa wartość data_salda: {val!r}. Użyj formatu YYYY-MM-DD lub DD.MM.YYYY.",
                )
        elif first == "numer_lokalu":
            header_row_idx = i
            break

    if global_date is None:
        raise HTTPException(
            status_code=422,
            detail="Brak wiersza data_salda. Dodaj w wierszu 1: | data_salda | 2024-12-31 |",
        )
    if header_row_idx is None:
        raise HTTPException(
            status_code=422,
            detail="Brak wiersza nagłówków (numer_lokalu, saldo_poczatkowe).",
        )

    headers = [_normalize_str(h).lower() for h in all_rows[header_row_idx]]
    if "saldo_poczatkowe" not in headers:
        raise HTTPException(status_code=422, detail="Brakująca kolumna: saldo_poczatkowe")

    col_idx = {name: headers.index(name) for name in ("numer_lokalu", "saldo_poczatkowe") if name in headers}
    data_rows = all_rows[header_row_idx + 1:]

    return global_date, col_idx, data_rows, header_row_idx


# ──────────────────────────────────────────────
# Endpoints
# ──────────────────────────────────────────────

@router.get("/template")
async def download_template(
    _admin: dict = Depends(require_admin),
):
    """Pobierz szablon Excel dla importu stanu początkowego."""
    xlsx_bytes = _build_template_xlsx()
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=szablon_stan_poczatkowy.xlsx"},
    )


@router.post("/initial-state", response_model=ImportInitialStateResult)
async def import_initial_state(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=True, description="true = tylko podgląd (bez zmian w bazie)"),
    _admin: dict = Depends(require_admin),
):
    """
    Importuj stan początkowy (saldo) lokali z pliku Excel.

    Format pliku:
    - Wiersz 1: | data_salda | <data> | (jedna globalna data dla wszystkich lokali)
    - Wiersz 2: | numer_lokalu | saldo_poczatkowe | (nagłówki)
    - Wiersze 3+: dane lokali

    - Aktualizuje tylko istniejące lokale (po numerze); nieznane numery → skipped.
    - W jednej komórce można podać wiele lokali (np. „18,31,42” lub „25.26”) — to samo saldo dla każdego.
    - dry_run=true (domyślnie): zwraca podgląd bez zapisu do bazy.
    """
    sb = get_supabase()

    # Walidacja rozszerzenia i rozmiaru
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Plik musi być w formacie .xlsx")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise HTTPException(status_code=413, detail="Plik zbyt duży (maks. 5 MB)")

    # Parsowanie struktury pliku
    global_date, col_idx, data_rows, header_row_idx = _parse_workbook(content)

    # Wczytaj wszystkie lokale z bazy (number → id)
    apts_res = sb.table("apartments").select("id, number").execute()
    apartments_by_number: dict[str, str] = {a["number"]: a["id"] for a in (apts_res.data or [])}

    # ── Przetwarzanie wierszy ──
    row_results: list[ImportRowResult] = []
    updates: dict[str, dict] = {}  # apartment_id → payload

    for idx, row in enumerate(data_rows):
        row_num = header_row_idx + 2 + idx
        nr_idx = col_idx.get("numer_lokalu")
        saldo_idx = col_idx.get("saldo_poczatkowe")

        nr_raw = row[nr_idx] if nr_idx is not None and nr_idx < len(row) else None
        nr_cell = _cell_to_apartment_text(nr_raw)
        if not nr_cell:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number="", status="skipped",
                message="Pusta komórka numer_lokalu",
            ))
            continue

        # Jedna pozycja w bazie: np. „3,4A”, „25,26” (zbiorczy zapis) — bez rozbijania na części
        if nr_cell in apartments_by_number:
            numbers = [nr_cell]
        else:
            numbers = _parse_apartment_numbers(nr_cell)
        if not numbers:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="skipped",
                message="Pusta komórka numer_lokalu",
            ))
            continue

        saldo_raw = row[saldo_idx] if saldo_idx is not None and saldo_idx < len(row) else None
        saldo = _parse_decimal(saldo_raw)
        if saldo is None:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="error",
                message=f"Nieprawidłowa wartość saldo_poczatkowe: {saldo_raw!r}",
            ))
            continue

        payload = {
            "initial_balance": float(saldo.quantize(Decimal("0.01"))),
            "initial_balance_date": global_date.isoformat(),
        }

        missing: list[str] = []
        found_any = False
        seen_apt_ids: set[str] = set()

        for n in numbers:
            apt_id = apartments_by_number.get(n)
            if apt_id is None:
                missing.append(n)
                continue
            found_any = True
            if apt_id not in seen_apt_ids:
                updates[apt_id] = payload
                seen_apt_ids.add(apt_id)

        if not found_any:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="skipped",
                message="Lokal nie istnieje w systemie",
            ))
            continue

        extra_msg = None
        if missing:
            extra_msg = "Brak lokali w systemie: " + ", ".join(missing)

        row_results.append(ImportRowResult(
            row=row_num,
            apartment_number=nr_cell,
            status="updated",
            message=extra_msg,
        ))

    # ── Zapis do bazy (jeśli nie dry_run) ──
    if not dry_run:
        for apt_id, payload in updates.items():
            sb.table("apartments").update(payload).eq("id", apt_id).execute()

    # ── Podsumowanie ──
    return ImportInitialStateResult(
        dry_run=dry_run,
        rows_total=len(row_results),
        updated=sum(1 for r in row_results if r.status == "updated"),
        skipped=sum(1 for r in row_results if r.status == "skipped"),
        errors=sum(1 for r in row_results if r.status == "error"),
        rows=row_results,
    )


# ──────────────────────────────────────────────
# Import wpłat (arkusz Dopasowania: Lokal, Data wpłaty, Kwota)
# Kolumna Nazwisko (jeśli wystąpi w pliku) jest ignorowana.
# ──────────────────────────────────────────────


def _normalize_header_slug(h: str) -> str:
    if not h:
        return ""
    s = str(h).strip().lower()
    # NFKD nie rozkłada „ł” — mapowanie polskich znaków przed unifikacją
    pl = str.maketrans("ąćęłńóśźż", "acelnoszz")
    s = s.translate(pl)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.replace(" ", "_").replace("-", "_")


def _resolve_payment_import_columns(headers: list) -> dict[str, int]:
    """
    Wymagane: lokal, data_wplaty, kwota.
    Dodatkowe kolumny (np. nazwisko) są ignorowane.
    """
    slug_to_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        slug = _normalize_header_slug(_normalize_str(h))
        if not slug:
            continue
        if slug not in slug_to_idx:
            slug_to_idx[slug] = i

    out: dict[str, int] = {}
    for slug in ("lokal", "numer_lokalu"):
        if slug in slug_to_idx:
            out["lokal"] = slug_to_idx[slug]
            break
    for slug in ("data_wplaty", "data_platnosci"):
        if slug in slug_to_idx:
            out["data_wplaty"] = slug_to_idx[slug]
            break
    if "kwota" in slug_to_idx:
        out["kwota"] = slug_to_idx["kwota"]
    return out


def _parse_payments_workbook(content: bytes) -> tuple[dict[str, int], list, int]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Biblioteka openpyxl nie jest zainstalowana")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
        sheet_name = "Dopasowania" if "Dopasowania" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Nie można otworzyć pliku Excel: {e}")

    if not all_rows:
        raise HTTPException(status_code=422, detail="Plik jest pusty")

    header_row_idx = 0
    headers = list(all_rows[header_row_idx])
    col_map = _resolve_payment_import_columns(headers)
    missing = [k for k in ("lokal", "data_wplaty", "kwota") if k not in col_map]
    if missing:
        raise HTTPException(
            status_code=422,
            detail="Wymagane kolumny: Lokal, Data wpłaty, Kwota (kolumna Nazwisko jest opcjonalna i ignorowana).",
        )

    data_rows = all_rows[header_row_idx + 1:]
    return col_map, data_rows, header_row_idx


def _build_payments_template_xlsx() -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="Biblioteka openpyxl nie jest zainstalowana")

    wb = Workbook()
    ws = wb.active
    ws.title = "Dopasowania"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(["Lokal", "Data wpłaty", "Kwota"], start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    examples = [
        ("1", "08.01.2026", "139,20"),
        ("3,4A", "10.01.2026", "500,00"),
    ]
    for row_idx, (loc, dt, kw) in enumerate(examples, start=2):
        ws.cell(row=row_idx, column=1, value=loc)
        ws.cell(row=row_idx, column=2, value=dt)
        ws.cell(row=row_idx, column=3, value=kw)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14

    ws.cell(row=6, column=1, value="Uwagi:").font = Font(bold=True)
    ws.cell(row=7, column=1, value="• Import uwzględnia wyłącznie: numer lokalu, datę wpłaty i kwotę.")
    ws.cell(row=8, column=1, value="• Dopasowanie numerów jak przy imporcie stanu początkowego (lokale zbiorcze, lista po przecinku).")
    ws.cell(row=9, column=1, value="• Wiele lokali w jednym wierszu: wpłata nadrzędna + rozbicie (proporcje z naliczeń w miesiącu daty lub równo).")
    ws.cell(row=10, column=1, value="• Wiele dat w jednej komórce (średnik), np. 10.02.2026; 27.02.2026 — osobna wpłata na każdą datę; jedna kwota = ta sama kwota przy każdej dacie; „341,20; 341,20” gdy kwoty się różnią.")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _parse_dates_list(raw) -> list[date] | None:
    """
    Jedna lub wiele dat w komórce (np. skrypt łączy dni operacji średnikiem).
    Separatory między datami: średnik, pionowa kreska, nowa linia.
    """
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return [raw.date()]
    if isinstance(raw, date):
        return [raw]
    s = str(raw).strip()
    if not s:
        return None
    parts = re.split(r"[;|\n\r]+", s)
    out: list[date] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        d = _parse_date(p)
        if d is None:
            return None
        out.append(d)
    return out if out else None


def _parse_amounts_for_payment_dates(kwota_raw, n_dates: int) -> tuple[list[Decimal], str | None] | None:
    """
    Zwraca listę kwot (długość n_dates) i opcjonalny komunikat pomocniczy.

    - Wiele kwot: „341,20; 341,20” — musi być tyle segmentów co dat.
    - Jedna kwota przy N datach: **ta sama kwota** na każdą datę (np. kilka wpłat po tej samej kwocie,
      różne dni księgowania z wyciągu).
    """
    if n_dates < 1:
        return None
    if kwota_raw is None:
        return None
    kw_s = str(kwota_raw).strip()
    if not kw_s:
        return None

    if ";" in kw_s:
        parts = [p.strip() for p in kw_s.split(";") if p.strip()]
        amts: list[Decimal] = []
        for p in parts:
            a = _parse_decimal(p)
            if a is None:
                return None
            if a <= 0:
                return None
            amts.append(a)
        if len(amts) == n_dates:
            return amts, None
        return None

    single = _parse_decimal(kwota_raw)
    if single is None or single <= 0:
        return None
    return [single] * n_dates, None


def _charges_for_split_month(sb, apartment_ids: list[str], payment_date: date) -> dict[str, Decimal]:
    split_month = payment_date.strftime("%Y-%m-01")
    charges_res = (
        sb.table("charges")
        .select("apartment_id, amount")
        .in_("apartment_id", apartment_ids)
        .eq("month", split_month)
        .execute()
    )
    apt_charges: dict[str, Decimal] = {aid: Decimal("0") for aid in apartment_ids}
    for c in (charges_res.data or []):
        apt_charges[c["apartment_id"]] += Decimal(str(c["amount"]))
    return apt_charges


@router.get("/payments-template")
async def download_payments_template(
    _admin: dict = Depends(require_admin),
):
    """Szablon Excel: Lokal, Data wpłaty, Kwota (bez nazwiska)."""
    xlsx_bytes = _build_payments_template_xlsx()
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=szablon_import_wplat.xlsx"},
    )


@router.post("/payments", response_model=ImportPaymentsResult)
async def import_payments(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=True, description="true = tylko podgląd (bez zapisu wpłat)"),
    _admin: dict = Depends(require_admin),
):
    """
    Import wpłat z arkusza (np. „Dopasowania”).

    Wymagane kolumny: Lokal, Data wpłaty, Kwota. Inne kolumny (np. Nazwisko) są ignorowane.

    Deduplikacja (jak import zestawienia bankowego): jeśli dla lokalu istnieje już wpłata
    z tą samą datą księgowania, nowa jest pomijana. Przy imporcie zbiorczym (wiele lokali)
    pominięty jest cały wiersz, gdy którykolwiek z lokali ma już wpłatę w tym dniu.
    Zbiór (lokal, data) jest też aktualizowany w trakcie przetwarzania pliku (podgląd
    i zapis są spójne z kolejnością wierszy).
    """
    sb = get_supabase()

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Plik musi być w formacie .xlsx")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise HTTPException(status_code=413, detail="Plik zbyt duży (maks. 5 MB)")

    col_map, data_rows, header_row_idx = _parse_payments_workbook(content)

    apts_res = sb.table("apartments").select("id, number, billing_group_id").execute()
    apartments_by_number: dict[str, dict] = {
        a["number"]: {"id": a["id"], "number": a["number"], "billing_group_id": a.get("billing_group_id")}
        for a in (apts_res.data or [])
    }

    all_apt_ids = [a["id"] for a in apartments_by_number.values()]
    existing_apt_dates: set[tuple[str, str]] = set()
    if all_apt_ids:
        existing_res = (
            sb.table("payments")
            .select("apartment_id, payment_date")
            .in_("apartment_id", all_apt_ids)
            .execute()
        )
        for p in (existing_res.data or []):
            aid = p.get("apartment_id")
            if aid is not None:
                existing_apt_dates.add((str(aid), str(p["payment_date"])))

    idx_lokal = col_map["lokal"]

    row_results: list[ImportRowResult] = []

    for idx, row in enumerate(data_rows):
        row_num = header_row_idx + 2 + idx

        def cell(col_key: str):
            j = col_map[col_key]
            return row[j] if j < len(row) else None

        nr_raw = row[idx_lokal] if idx_lokal < len(row) else None
        nr_cell = _cell_to_apartment_text(nr_raw)
        if not nr_cell:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number="", status="skipped",
                message="Pusta komórka Lokal",
            ))
            continue

        pay_date_raw = cell("data_wplaty")
        kwota_raw = cell("kwota")

        if pay_date_raw is None or (isinstance(pay_date_raw, str) and not str(pay_date_raw).strip()):
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="skipped",
                message="Brak daty wpłaty",
            ))
            continue
        if kwota_raw is None:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="skipped",
                message="Brak kwoty",
            ))
            continue

        pay_dates = _parse_dates_list(pay_date_raw)
        if pay_dates is None:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="error",
                message=f"Nieprawidłowa data wpłaty: {pay_date_raw!r}",
            ))
            continue

        if nr_cell in apartments_by_number:
            numbers = [nr_cell]
        else:
            numbers = _parse_apartment_numbers(nr_cell)
        if not numbers:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="skipped",
                message="Pusta komórka Lokal",
            ))
            continue

        resolved: list[dict] = []
        missing_nums: list[str] = []
        for n in numbers:
            apt = apartments_by_number.get(n)
            if apt is None:
                missing_nums.append(n)
            else:
                resolved.append(apt)

        if missing_nums:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="error",
                message="Lokal nie istnieje w systemie: " + ", ".join(missing_nums),
            ))
            continue
        if not resolved:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="skipped",
                message="Brak dopasowanych lokali",
            ))
            continue

        if len(pay_dates) > 1 and len(resolved) > 1:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="error",
                message="Wiele dat w jednym wierszu obsługiwane tylko dla pojedynczego lokalu — podziel wiersze lub użyj jednej daty.",
            ))
            continue

        amounts_parsed = _parse_amounts_for_payment_dates(kwota_raw, len(pay_dates))
        if amounts_parsed is None:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="error",
                message=(
                    f"Nieprawidłowa kwota dla {len(pay_dates)} dat: {kwota_raw!r}. "
                    f"Podaj jedną kwotę (powtórzy się na każdą datę) albo N kwot oddzielonych średnikiem."
                ),
            ))
            continue
        amounts, _amount_hint = amounts_parsed

        resolved.sort(key=lambda a: a["number"])
        apt_ids = [a["id"] for a in resolved]

        if len(resolved) == 1:
            a0 = resolved[0]
            n_pay = len(pay_dates)
            for pi, (pay_date, amount) in enumerate(zip(pay_dates, amounts)):
                date_str = pay_date.isoformat()
                key = (str(a0["id"]), date_str)
                if key in existing_apt_dates:
                    dup_msg = f"Duplikat — wpłata na lokal {nr_cell} z dnia {date_str} już istnieje"
                    if n_pay > 1:
                        dup_msg += f" ({pi + 1}/{n_pay})"
                    row_results.append(ImportRowResult(
                        row=row_num, apartment_number=nr_cell, status="skipped",
                        message=dup_msg,
                    ))
                    continue
                if not dry_run:
                    sb.table("payments").insert({
                        "apartment_id": a0["id"],
                        "billing_group_id": a0.get("billing_group_id"),
                        "amount": str(amount.quantize(Decimal("0.01"))),
                        "payment_date": date_str,
                        "title": "Wpłata z dnia",
                        "confirmed_by_admin": True,
                        "matched_automatically": False,
                    }).execute()
                existing_apt_dates.add(key)
                msg = (
                    f"Wpłata z dnia — {date_str} ({pi + 1}/{n_pay})"
                    if n_pay > 1
                    else None
                )
                row_results.append(ImportRowResult(
                    row=row_num, apartment_number=nr_cell, status="updated",
                    message=msg,
                ))
            continue

        pay_date = pay_dates[0]
        amount = sum(amounts)
        date_str = pay_date.isoformat()

        has_existing = any(
            (str(r["id"]), date_str) in existing_apt_dates
            for r in resolved
        )
        if has_existing:
            row_results.append(ImportRowResult(
                row=row_num, apartment_number=nr_cell, status="skipped",
                message=(
                    f"Duplikat — wpłata zbiorcza z dnia {date_str} koliduje z istniejącą "
                    "wpłatą na jednym z lokali"
                ),
            ))
            continue

        g0 = resolved[0].get("billing_group_id")
        common_group = (
            g0
            if g0 is not None and all(a.get("billing_group_id") == g0 for a in resolved)
            else None
        )

        apt_charges = _charges_for_split_month(sb, apt_ids, pay_date)
        if common_group is not None:
            split_map = compute_split_amounts(apt_ids, apt_charges, amount)
        else:
            split_map = compute_split_amounts(
                apt_ids, {aid: Decimal("0") for aid in apt_ids}, amount
            )

        apt_number_map = {a["id"]: a["number"] for a in resolved}

        if not dry_run:
            parent_data = {
                "apartment_id": None,
                "billing_group_id": common_group,
                "amount": str(amount.quantize(Decimal("0.01"))),
                "payment_date": date_str,
                "title": "Import zbiorczy",
                "confirmed_by_admin": True,
                "matched_automatically": False,
            }
            parent_res = sb.table("payments").insert(parent_data).execute()
            parent_id = parent_res.data[0]["id"]

            for aid, amt in split_map.items():
                if amt == 0:
                    continue
                sb.table("payments").insert({
                    "apartment_id": aid,
                    "billing_group_id": next(
                        (x.get("billing_group_id") for x in resolved if x["id"] == aid), None
                    ),
                    "parent_payment_id": parent_id,
                    "amount": str(amt.quantize(Decimal("0.01"))),
                    "payment_date": date_str,
                    "title": f"Rozbicie wpłaty - lokal {apt_number_map[aid]}",
                    "confirmed_by_admin": True,
                    "matched_automatically": True,
                }).execute()

        for aid, amt in split_map.items():
            if amt == 0:
                continue
            existing_apt_dates.add((str(aid), date_str))

        row_results.append(ImportRowResult(
            row=row_num, apartment_number=nr_cell, status="updated",
            message="Wpłata nadrzędna + rozbicie na lokale" if not dry_run else "Symulacja: wpłata nadrzędna + rozbicie",
        ))

    return ImportPaymentsResult(
        dry_run=dry_run,
        rows_total=len(row_results),
        updated=sum(1 for r in row_results if r.status == "updated"),
        skipped=sum(1 for r in row_results if r.status == "skipped"),
        errors=sum(1 for r in row_results if r.status == "error"),
        rows=row_results,
    )


# ──────────────────────────────────────────────
# Import z zestawienia bankowego (.xls)
# ──────────────────────────────────────────────


def _build_registry(sb) -> list[ApartmentRecord]:
    """Pobiera rejestr lokali z bazy do dopasowania transakcji."""
    res = sb.table("apartments").select(
        "id, number, billing_surname, billing_group_id"
    ).execute()
    registry: list[ApartmentRecord] = []
    for apt in (res.data or []):
        registry.append(ApartmentRecord(
            apartment_id=apt["id"],
            number=apt["number"],
            billing_surname=apt.get("billing_surname"),
            billing_group_id=apt.get("billing_group_id"),
        ))
    return registry


@router.post("/payments-bank-statement", response_model=ImportBankStatementResult)
async def import_bank_statement(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=True, description="true = tylko podgląd (bez zapisu wpłat)"),
    _admin: dict = Depends(require_admin),
):
    """
    Import wpłat z zestawienia bankowego (.xls).

    Automatycznie dopasowuje transakcje do lokali na podstawie:
    - numeru lokalu w opisie/adresie przelewu
    - nazwiska rozliczeniowego (billing_surname) z rejestru lokali

    Tylko wpłaty (kwota > 0) są importowane. Transakcje bez jednoznacznego
    dopasowania trafiają na listę „niedopasowanych" w odpowiedzi.
    """
    sb = get_supabase()

    filename = file.filename or ""
    if not filename.lower().endswith(".xls"):
        raise HTTPException(
            status_code=422,
            detail="Plik musi być w formacie .xls (stary format Excel). Dla plików .xlsx użyj importu wpłat z arkusza Dopasowania.",
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise HTTPException(status_code=413, detail="Plik zbyt duży (maks. 5 MB)")

    registry = _build_registry(sb)
    if not registry:
        raise HTTPException(
            status_code=422,
            detail="Brak lokali w systemie. Dodaj lokale przed importem zestawienia.",
        )

    surnames_configured = sum(1 for r in registry if r.billing_surname)
    if surnames_configured == 0:
        raise HTTPException(
            status_code=422,
            detail="Żaden lokal nie ma uzupełnionego nazwiska rozliczeniowego (billing_surname). "
                   "Uzupełnij nazwiska w panelu Lokale przed importem zestawienia bankowego.",
        )

    try:
        parse_result = parse_bank_statement(content, registry)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    BANK_TITLE = "Wpłata z zestawienia bankowego"

    # Lookup lokali po numerze (do rozwiązywania splitów)
    apts_res = sb.table("apartments").select("id, number, billing_group_id").execute()
    apt_by_number: dict[str, dict] = {
        a["number"]: {"id": a["id"], "number": a["number"], "billing_group_id": a.get("billing_group_id")}
        for a in (apts_res.data or [])
    }

    # Rozwiąż każde dopasowanie na listę lokali:
    # 1. Parser zwrócił group_records (wiele lokali w jednej grupie rozliczeniowej) → split
    # 2. Numer zbiorczy "25,26" → rozbij na składowe i sprawdź w bazie
    # 3. Pojedynczy lokal → bez splitu
    resolved_matches: list[tuple] = []  # (MatchedPayment, list[dict])
    all_apt_ids: set[str] = set()

    for m in parse_result.matched:
        resolved: list[dict] = []

        # Priorytet 1: parser wykrył grupę rozliczeniową
        if m.group_records and len(m.group_records) > 1:
            resolved = [
                {"id": gr.apartment_id, "number": gr.number,
                 "billing_group_id": gr.billing_group_id}
                for gr in m.group_records
            ]
        else:
            # Priorytet 2: rozbij numer zbiorczy na składowe
            if m.apartment_number in apt_by_number:
                individual_numbers = [m.apartment_number]
            else:
                individual_numbers = _parse_apartment_numbers(m.apartment_number)

            resolved = [apt_by_number[n] for n in individual_numbers if n in apt_by_number]

        if not resolved:
            # Fallback: użyj dopasowanego rekordu
            resolved = [{"id": m.apartment_id, "number": m.apartment_number,
                         "billing_group_id": m.billing_group_id}]

        resolved_matches.append((m, resolved))
        for r in resolved:
            all_apt_ids.add(r["id"])

    # Deduplikacja: sprawdź istniejące wpłaty po (apartment_id, date)
    existing_apt_dates: set[tuple[str, str]] = set()
    if all_apt_ids:
        existing_res = (
            sb.table("payments")
            .select("apartment_id, payment_date")
            .in_("apartment_id", list(all_apt_ids))
            .execute()
        )
        for p in (existing_res.data or []):
            existing_apt_dates.add(
                (p["apartment_id"], str(p["payment_date"]))
            )

    # Podział na nowe vs duplikaty
    # Sprawdzanie po (apartment_id, date) — bez kwoty, bo kwoty rozbite
    # różnią się od kwoty głównej, a dwie wpłaty na ten sam lokal w tym
    # samym dniu to rzadkość (lepiej oznaczyć jako duplikat niż zaksięgować podwójnie)
    new_matched: list[tuple] = []
    skipped_duplicates: list = []
    for m, resolved in resolved_matches:
        date_str = m.payment_date.isoformat()
        has_existing = any(
            (r["id"], date_str) in existing_apt_dates
            for r in resolved
        )
        if has_existing:
            skipped_duplicates.append(m)
            continue

        new_matched.append((m, resolved))

    # Zapis — single lub parent + children (split)
    if not dry_run:
        for m, resolved in new_matched:
            resolved.sort(key=lambda a: a["number"])
            apt_ids = [a["id"] for a in resolved]

            if len(resolved) == 1:
                # Pojedynczy lokal
                a0 = resolved[0]
                sb.table("payments").insert({
                    "apartment_id": a0["id"],
                    "billing_group_id": a0.get("billing_group_id"),
                    "amount": str(m.amount),
                    "payment_date": m.payment_date.isoformat(),
                    "title": BANK_TITLE,
                    "confirmed_by_admin": True,
                    "matched_automatically": True,
                }).execute()
            else:
                # Wiele lokali — parent + rozbicie (jak w import_payments)
                g0 = resolved[0].get("billing_group_id")
                common_group = (
                    g0
                    if g0 is not None and all(a.get("billing_group_id") == g0 for a in resolved)
                    else None
                )

                apt_charges = _charges_for_split_month(sb, apt_ids, m.payment_date)
                if common_group is not None:
                    split_map = compute_split_amounts(apt_ids, apt_charges, m.amount)
                else:
                    split_map = compute_split_amounts(
                        apt_ids, {aid: Decimal("0") for aid in apt_ids}, m.amount
                    )

                parent_data = {
                    "apartment_id": None,
                    "billing_group_id": common_group,
                    "amount": str(m.amount),
                    "payment_date": m.payment_date.isoformat(),
                    "title": BANK_TITLE,
                    "confirmed_by_admin": True,
                    "matched_automatically": True,
                }
                parent_res = sb.table("payments").insert(parent_data).execute()
                parent_id = parent_res.data[0]["id"]

                apt_number_map = {a["id"]: a["number"] for a in resolved}
                for aid, amt in split_map.items():
                    if amt == 0:
                        continue
                    sb.table("payments").insert({
                        "apartment_id": aid,
                        "billing_group_id": next(
                            (x.get("billing_group_id") for x in resolved if x["id"] == aid), None
                        ),
                        "parent_payment_id": parent_id,
                        "amount": str(amt.quantize(Decimal("0.01"))),
                        "payment_date": m.payment_date.isoformat(),
                        "title": f"Rozbicie wpłaty - lokal {apt_number_map[aid]}",
                        "confirmed_by_admin": True,
                        "matched_automatically": True,
                    }).execute()

    # Budowanie odpowiedzi
    matched_rows = [
        BankStatementMatchedRow(
            apartment_number=m.apartment_number,
            payment_date=m.payment_date.isoformat(),
            amount=str(m.amount),
            confidence=round(m.match_confidence, 2),
            match_details=m.match_details
            + (f" → rozbicie na {len(resolved)} lokali" if len(resolved) > 1 else ""),
        )
        for m, resolved in new_matched
    ]

    # Duplikaty jako dodatkowe unmatched z jasnym powodem
    for m in skipped_duplicates:
        parse_result.unmatched.append(UnmatchedTransaction(
            row_index=0,
            payment_date=m.payment_date,
            amount=m.amount,
            sender_name="",
            description="",
            reason=f"Duplikat — wpłata {m.amount} zł na lokal {m.apartment_number} z dnia {m.payment_date.isoformat()} już istnieje",
        ))

    unmatched_rows = [
        BankStatementUnmatchedRow(
            row_index=u.row_index,
            payment_date=u.payment_date.isoformat() if u.payment_date else None,
            amount=str(u.amount) if u.amount else None,
            sender_name=u.sender_name[:100],  # RODO: obcinamy długie ciągi
            description=u.description[:100],
            reason=u.reason,
        )
        for u in parse_result.unmatched
    ]

    return ImportBankStatementResult(
        dry_run=dry_run,
        total_rows=parse_result.total_rows,
        matched_count=len(matched_rows),
        unmatched_count=len(unmatched_rows),
        matched=matched_rows,
        unmatched=unmatched_rows,
    )
